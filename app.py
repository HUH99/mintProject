# Description: 봇의 메인 기능을 구현한 코드. 수요예측 의견을 전송하고 참여내역을 확인하는 기능을 구현함.
# MAIN 함수
import json
import pandas as pd
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, CallbackQueryHandler, filters, ContextTypes
from dotenv import load_dotenv
import os
import logging
import asyncio
from asyncio import Lock, Event
import aiohttp
import aiofiles
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from message_tracker import MessageTracker
from typing import Dict, List, Tuple, Optional

# 봇 토큰 로드
load_dotenv() # .env 파일에서 환경 변수를 로드하여 os.getenv()로 사용할 수 있게 함
TOKEN = os.getenv('TELEGRAM_BOT_TOKEN')
EXCEL_FOLDER = 'excel_files'

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# 엑셀 파일 읽기 함수
def read_advise_excel(file_name):
    """수요예측 엑셀 파일의 두 개의 시트: 수요예측 첫날 의견과 마지막날 의견을 읽어서 data frame으로 반환합니다.

    Args:
        file_name ( .xlsx file ): 파일이름이 수요예측하는 종목 이름인 엑셀 파일 

    Returns:
        _type_: 첫날에 대한 dataframe과 마지막날에 대한 dataframe
    """
    
    try:
        full_path = os.path.join(EXCEL_FOLDER, file_name)
        logger.info(f"{file_name} 파일을 읽습니다.")
        df1 = pd.read_excel(io=full_path, sheet_name='firstDay', usecols='A:I', index_col=0,
                            engine='openpyxl', dtype={'참여가격': str})
        df2 = pd.read_excel(io=full_path, sheet_name='lastDay', usecols='A:I', index_col=0,
                            engine='openpyxl', dtype={'참여가격': str})

        # '참여가격' 열의 형식을 조정
        for df in [df1, df2]:
            df['참여가격'] = df['참여가격'].apply(format_price)

        logger.info(f"{file_name} 파일을 성공적으로 읽었습니다.")
        return df1, df2
    except FileNotFoundError:
        logger.error(f"{file_name} 파일을 찾을 수 없습니다.")
        raise
    except PermissionError:
        logger.error(f"{file_name} 파일에 접근할 권한이 없습니다.")
        raise
    except Exception as e:
        logger.error(f"{file_name} 파일을 읽는 중 오류가 발생했습니다: {str(e)}")
        raise

# '참여가격' 열의 형식을 조정하는 함수
def format_price(price):
    """
    가격을 포맷팅하는 함수
    """
    try:
        # 문자열에서 쉼표 제거 후 숫자로 변환
        numeric_price = float(str(price).replace(',', ''))
        # 천 단위 쉼표를 포함한 문자열로 변환
        return f"{numeric_price:,.0f}"
    except ValueError:
        # 숫자로 변환할 수 없는 경우 원래 값을 반환
        return price

# config.json 파일 읽기 함수
async def read_config():
    try:
        async with aiofiles.open('config.json', mode='r', encoding="UTF-8") as f:
            content = await f.read()
            return json.loads(content)
    except FileNotFoundError:
        logger.error("config.json 파일을 찾을 수 없습니다.")
        raise
    except json.JSONDecodeError:
        logger.error("config.json 파일의 형식이 잘못되어 읽을 수 없습니다.")
        raise

# config.json 파일 업데이트 함수 추가
async def update_config(config):
    try:    
        async with aiofiles.open('config.json', mode='w') as f:
            await f.write(json.dumps(config, indent=4))
        logger.info("config.json 파일을 업데이트했습니다.")
    except IOError:
        logger.error("config.json 파일을 업데이트하는 중 오류가 발생했습니다.")
        raise


# 각 종목별 수요예측 프로세스를 관리하는 StockAdvisory 클래스
class StockAdvisory:
    def __init__(self, stock_name: str, df: pd.DataFrame, sheet_name: str, application: Application):
        self.stock_name = stock_name
        self.df = df
        self.sheet_name = sheet_name
        self.application = application
        self.message_trackers = {} # 고객사별 메시지 추적기
        self.config = None
    
    async def initialize(self):
        self.config = await read_config()

    async def process(self, application: Application):
        """전체 수요예측 프로세스를 관리. 수요예측 의견을 전송하는 작업을 생성합니다. 발송 여부가 1인 기관들에 대해 메시지를 전송합니다. 

        Args:
            application ( Application ): 실행히고 있는 Application 객체
        """
        await self.initialize()
        
        try:
            logger.info(f"\"{self.stock_name}\"의 수요예측 프로세스를 시작합니다.")
        
            comment = self.df['코멘트'].iloc[0] # 모든 행에 대해 동일한 코멘트 사용
            mint_group_chat_id = self.config['mint_group_chat_id']


            # Chat ID 업데이트가 필요한 고객사 목록 생성
            clients_needing_chat_id = []
            for index, row in self.df[self.df['발송'] == 1].iterrows():
                if pd.isna(row['chatID']) or (f"민트-{index}" in self.config["client_group_chat_id"] and 
                                            row['chatID'] != self.config["client_group_chat_id"][f"민트-{index}"]):
                    clients_needing_chat_id.append(index)

            # Chat ID가 필요한 고객사가 있으면 get_chat_ids 호출
            chat_ids_updated = False
            if clients_needing_chat_id:
                chat_ids_updated = True
                new_chat_ids = await self.get_chat_ids(application, clients_needing_chat_id)
                for client, chat_id in new_chat_ids.items():
                    if chat_id:
                        self.df.at[client, 'chatID'] = chat_id


            # 발송 여부가 1인 고객기관들에 대해 각각 task 생성하여 비동기로 처리
            tasks = [self.process_each_client(index, row, application, mint_group_chat_id, comment) for index, row in self.df[self.df['발송'] == 1].iterrows()]
            
            results = await asyncio.gather(*tasks, return_exceptions=True)  # asyncio.gather()를 통해 모든 task를 병렬로 동시에 실행. 결과는 참여의견 메시지 객체


            # MessageTracker 인스턴스 생성
            for index, row in self.df[self.df['발송'] == 1].iterrows():
                chat_id = row['chatID']
                if pd.notna(chat_id):
                    tracker = MessageTracker(self.application, index, int(chat_id))
                    await tracker.initialize()
                    self.message_trackers[index] = tracker

            # 참여의견 전송 실패 고객사 처리
            failed_clients = []
            
            # 의견을 발송해야 하는 고객사들의 인덱스와 process_each_client 메소드의 결과를 짝지어 순회
            for index, result in zip(self.df[self.df['발송'] == 1].index, results):
                # 결과가 예외 객체인 경우, 해당 고객사에 대한 처리가 실패했음을 의미
                if isinstance(result, Exception):                                    
                    logger.error(f"\"{self.stock_name}\" 참여의견을 \"{index}\"에 보내는 중 오류 발생: {str(result)}")
                    failed_clients.append(index)
                # 참여의견 전송에 성공한 고객사에 대해 메시지 추적 시작
                if index in self.message_trackers:
                    await self.message_trackers[index].start_tracking(self.stock_name, result)


            # chatID을 못 찾았거나 다른 오류로 메시지 전송에 실패했을 때 민트 실무진 그룹채팅방에 알림
            if failed_clients:
                sending_failed_message = "\n!!!!!!!!<긴급>!!!!!!!!\n".join([f" \"{self.stock_name}\" 참여의견을 \"{client}\"에 전송하지 못했습니다. 직접 보내주세요." for client in failed_clients])
                await application.bot.send_message(chat_id=mint_group_chat_id, text=sending_failed_message)


            # Chat ID가 업데이트 되었으면 수정된 DataFrame을 해당 종목의 엑셀 파일에 저장
            if chat_ids_updated:
                try:
                    await save_dataframe_to_excel(self.df, f"{self.stock_name}.xlsx", self.sheet_name)
                except Exception as e:
                    logger.error(f"\"{self.stock_name}\"의 DataFrame을 엑셀 파일에 저장하는 중 오류 발생: {str(e)}")


            # logger.info(f"\"{self.stock_name}\" 수요예측 프로세스 완료.")


        except Exception as e:
            logger.error(f"\"{self.stock_name}\" 처리 중 예외 발생: {str(e)}")
            raise


    async def process_each_client(self, index: str, row: pd.Series, application: Application, mint_group_chat_id: int, comment: str) -> Optional[int]:
        """특정 종목에 대해 StockAdvisory 클래스를 생성했으므로 각 개별 클라이언트에 대한 처리를 담당합니다. Chat ID를 확인하고, 메시지를 전송하는 등의 작업을 수행.

        Args:
            index ( DF의 인덱스): DataFrame의 인덱스. 여기서는 고객기관명
            row ( Series ): DataFrame의 특정 행 series
            application ( Application ): 실행하고 있는 Application 객체
            mint_group_chat_id ( int ): 민트 실무진 그룹 채팅방의 chat_id
            comment ( string ): 수요예측 의견의 코멘트
        
        Returns:
            Message: 참여의견 메시지 객체
        
        Raises:
            Exception: "민트-{index}" 그룹 Chat ID를 찾을 수 없는 경우 또는 메시지 전송 중 오류 발생 시
        """
        try:
            chat_id = row['chatID']

            if pd.isna(chat_id):
                logger.info(f"\"민트-{index}\" 그룹 Chat ID를 찾을 수 없습니다. config.json 파일을 확인하세요.")
                return None
            
            # DataFrame의 'chatID' 필드의 값이 config.json 파일 값과 일치하는 경우, 해당 Chat ID로 메시지 전송
            if chat_id == self.config["client_group_chat_id"][f"민트-{index}"]:
                sent_message = await self.send_advisory(application, chat_id, row, comment, index)
                if sent_message:
                    await application.bot.send_message(chat_id=mint_group_chat_id, text=f"\"{index}\"에 \"{self.stock_name}\" 참여의견 메시지를 전송했습니다.")
                    return sent_message
                else:
                    logger.error(f"\"{self.stock_name}\" 참여의견을 \"{index}\"에 전송하지 못했습니다.")
                    return None
            else:
                logger.error(f"{self.stock_name}.xlsx 에서 \"민트-{index}\" 그룹 Chat ID 가 config.json 파일과 일치하지 않습니다. 다시 한번 확인하세요.")
                return None
            
        except Exception as e:
            logger.error(f"\"{self.stock_name}\" 참여의견을 \"{index}\"에 보내는 중에 예기치 못한 오류 발생: {str(e)}")
            return None


    # 수요예측 의견 전송 함수
    async def send_advisory(self, application, chat_id, row, comment, client_name):
        """수요예측 참여의견 메시지를 전송합니다. 참여의견 전송 후 메시지 추적기를 통해 고객사의 답장 유무를 추적하고 참여내역을 확인합니다.

        Args:
            application ( Application ): 실행하고 있는 Application 객체
            chat_id (int): 메시지를 전송할 채팅방의 chat_id
            row (Series): 특정 고객기관의 수요예측 의견을 담고 있는 Series
            comment (string): 수요예측 의견의 코멘트
            client_name (string): 기관명
        
        Returns:
            Message: 전송된 참여의견 메시지 객체
        """

        message = f"""
        {comment}
        ———————————————————————————————
        참여가격: {row['참여가격']}원
        참여수량: {row['참여수량']}
        확약여부: {row['확약여부']}
        """
        try:
            sent_message = await application.bot.send_message(chat_id=chat_id, text=message)
            logger.info(f"\"{self.stock_name}\" 참여의견을 \"{client_name}\"에 성공적으로 전송했습니다.")
            return sent_message
            
        except Exception as e:
            logger.error(f"\"{self.stock_name}\" 참여의견을 \"{client_name}\"에 보내는 중 오류 발생: {str(e)}")
            return None

    # 엑셀에 없는 그룹의 Chat ID를 config.json 파일에서 가져오는 함수
    async def get_chat_ids(self, application, client_names: list):
        """ 엑셀 파일의 ChatID 필드가 비어있는 경우, 'config.json' 파일에서 그룹 Chat ID를 반환합니다.

        Args:
            application (Application): Application 객체. 필요한지는 모르겠음.
            client_names ( list ): 찾고자 하는 그룹 이름의 리스트 (엑셀 파일에 적힌 기관명과 텔레그램 그룹 채팅방 이름이 일치해야 함)

        Raises:
            Exception: API 요청 실패 시 발생

        Returns:
            int: Chat ID 읽기 성공 시 해당 그룹의 chat_id, 실패 시 None
        """
        
        chat_ids = {}

        for client in client_names:
            group_name = f"민트-{client}"
            if group_name in self.config['client_group_chat_id']:
                chat_ids[client] = self.config['client_group_chat_id'][group_name]
                logger.info(f"config.json 파일에서 \"{group_name}\" 그룹 Chat ID를 찾아서 반환합니다.")
            else:
                logger.error(f"config.json 파일에서 \"{group_name}\" 그룹 Chat ID를 찾을 수 없습니다.")
                application.bot.send_message(chat_id=self.config['mint_group_chat_id'], text=f"\"{group_name}\" 그룹 Chat ID를 찾을 수 없습니다. 직접 보내주세요.")

        return chat_ids

    async def stop(self):
        for tracker in self.message_trackers.values():
            await tracker.stop()

###################### 여기까지 StockAdvisory 클래스 끝 #########################


# 엑셀 파일 업데이트 함수
async def save_dataframe_to_excel(df, file_name, sheet_name):
    try:
        full_path = os.path.join(EXCEL_FOLDER, file_name)
        # 기존 엑셀 파일 로드
        book = load_workbook(full_path)
        
        # 해당 시트 선택 (없으면 새로 생성)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            logger.error(f"{file_name} 파일에 {sheet_name} 시트가 없습니다.")
        
        # DataFrame에서 Chat ID 열만 추출
        chat_id_df = df[['chatID']]

        # Chat ID 열만 업데이트
        for r_idx, row in enumerate(dataframe_to_rows(chat_id_df, index=True, header=False), 2): # chat_id_df를 행 단위로 변환.
            for c_idx, value in enumerate(row[1:], 6): # ChatID는 6번째 열(F열)에 있음.
                if pd.notna(value): # 빈 셀이나 누락된 데이터는 무시.
                    sheet.cell(row=r_idx, column=c_idx, value=value)
        
        # 변경사항 저장
        try:
                book.save(full_path)
                logger.info(f"{file_name} 파일의 {sheet_name} 시트에서 Chat ID를 업데이트했습니다.")
        except PermissionError:
            logger.error(f"{file_name} 파일에 저장할 권한이 없습니다. 파일이 열려있지 않은지 확인하세요.")
        except IOError as e:
            logger.error(f"{file_name} 파일 저장 중 IO 오류 발생: {str(e)}")     
        
    
    except Exception as e:
        logger.error(f"{file_name} 파일을 업데이트하는 중 오류가 발생했습니다: {str(e)}")
        raise



# 종목 참여내역 통합 처리 함수
async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    message = update.message
    chat_id = message.chat_id

    for stock_advisory in active_stocks.values():
        for tracker in stock_advisory.message_trackers.values():
            if tracker.chat_id == chat_id:
                await tracker.process_message(message)
                logger.info(f"")
                


# main 함수
async def main():
    config = await read_config()
    
    # 봇 생성
    application = Application.builder().token(TOKEN).pool_timeout(60).connection_pool_size(64).build()
    application.add_handler(MessageHandler(filters.ALL & ~filters.COMMAND, handle_message))

    await application.initialize()
    await application.start()

    # 폴링을 별도의 태스크로 진행
    # (사용자 입력 처리, StockAdvisory 인스턴스 생성 및 처리, 메인 루프 실행 등의 작업들이 진행되는 동안에도 봇은 계속해서 새로운 메시지나 업데이트를 받을 수 있음)
    polling_task = asyncio.create_task(application.updater.start_polling(allowed_updates=Update.ALL_TYPES))

    global active_stocks
    active_stocks = {}
    
    # await application.bot.send_message(chat_id=config['mint_group_chat_id'], text="프로그램이 시작되었습니다.")
    logger.info("프로그램이 시작되었습니다.")

    try:
        while True:
            # 사용자로부터 수요예측을 진행할 종목명을 입력받음
            stock_name = input("수요예측을 진행할 주식 종목명을 입력하세요 (종료하려면 'q' 입력): ")

            if stock_name.lower() == 'q':
                break

            sheet_name = input("수요예측 첫날이면 '1', 마지막날이면 '2'를 입력하세요: ").strip()
            
            if sheet_name not in ['1', '2']:
                logger.warning("잘못된 입력입니다. '1' 또는 '2'를 입력해주세요.")
                continue

            # '종목명.xlsx' 파일 읽기            
            file_name = f"{stock_name}.xlsx"
            try:
                df_first_day, df_last_day = read_advise_excel(file_name)
                df = df_first_day if sheet_name == '1' else df_last_day
                sheet_name = 'firstDay' if sheet_name == '1' else 'lastDay'
            except FileNotFoundError:
                print(f"{file_name} 파일이 존재하지 않습니다.")
                continue

            # StockAdvisory 인스턴스 생성 및 처리
            stock_advisory = StockAdvisory(stock_name, df, sheet_name, application)

            active_stocks[stock_name] = stock_advisory # 오늘 수요예측 하는 모든 종목의 StockAdvisory 인스턴스 저장

            # 비동기로 처리 시작
            try:
                await stock_advisory.process(application)
            except Exception as e:
                logger.error(f"{stock_name} 처리 중 오류 발생: {str(e)}")
            
            print(f"{stock_name}의 수요예측 참여의견 전송 작업이 완료되었습니다.")

            print("다음 종목의 수요예측 의견을 전송하려면 새로운 종목명을 입력하세요.")
            print("모든 처리를 마치고 프로그램을 종료하려면 'q'를 입력하세요.")

    finally:
        # 애플리케이션 종료
        await application.stop()

        # 폴링 중지
        if application.updater and application.updater.running:
            await application.updater.stop()

        # 폴링 태스크가 완전히 종료될 때까지 대기
        if polling_task and not polling_task.done():
            try:
                await asyncio.wait_for(polling_task, timeout=10)
            except asyncio.TimeoutError:
                logging.warning("폴링 태스크가 시간 안에 완료되지 않았습니다.")

        # 모든 실행 중인 테스크 취소
        for task in asyncio.all_tasks():
            if task is not asyncio.current_task():
                task.cancel()
        
        # 취소된 태스크들이 정리될 때까지 대기
        await asyncio.gather(*asyncio.all_tasks(), return_exceptions=True)

    print("프로그램을 종료합니다.")

if __name__ == "__main__":
    asyncio.run(main())